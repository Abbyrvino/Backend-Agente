from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.utils import get_column_letter


def generar_pdf(datos: list[dict], nombre_archivo: str = "reporte.pdf"):
    """
    Genera un archivo PDF a partir de una lista de diccionarios.

    Args:
        datos (list[dict]): Una lista de diccionarios con los datos.
        nombre_archivo (str, optional): El nombre del archivo PDF a generar.
    """
    if not datos:
        print("No hay datos para generar el PDF.")
        return

    doc = SimpleDocTemplate(nombre_archivo, pagesize=letter)
    elementos = []

    # Extraer encabezados y datos
    encabezados = list(datos[0].keys())
    data_para_tabla = [encabezados] + [list(d.values()) for d in datos]

    # Crear la tabla
    tabla = Table(data_para_tabla)

    # Estilo de la tabla
    estilo = TableStyle(
        [
            ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ]
    )
    tabla.setStyle(estilo)

    elementos.append(tabla)
    doc.build(elementos)
    print(f"Archivo PDF '{nombre_archivo}' generado exitosamente.")


def generar_excel(datos: list[dict], nombre_archivo: str = "reporte.xlsx"):
    """
    Genera un archivo Excel a partir de una lista de diccionarios.

    Args:
        datos (list[dict]): Una lista de diccionarios con los datos.
        nombre_archivo (str, optional): El nombre del archivo Excel a generar.
    """
    if not datos:
        print("No hay datos para generar el Excel.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # Escribir encabezados
    encabezados = list(datos[0].keys())
    ws.append(encabezados)

    # Escribir datos
    for fila in datos:
        ws.append(list(fila.values()))

    # Autoajustar el ancho de las columnas
    for i, columna in enumerate(encabezados, 1):
        letra_columna = get_column_letter(i)
        ws.column_dimensions[letra_columna].autosize = True

    wb.save(nombre_archivo)
    print(f"Archivo Excel '{nombre_archivo}' generado exitosamente.")
